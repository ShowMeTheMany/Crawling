import time

from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException

# from openpyxl import Workbook
import os
import openpyxl
import winsound as sd
import csv

### 작성 구역 START ####--------------------------------------------------
crawling_flag = False  # True는 크롤링 작업 False는 병합해서 csv로 변환하는 작업

big_cate = '문구/오피스'                     # 엑셀 파일 내부 대분류
big_cate_fileName = '문구.오피스'            # 폴더이름 또는 엑셀파일 시트이름
middle_cate = '캐릭터 문구'                  # 엑셀 파일 내부 소분류

# 소분류 기준으로 이름과 category 코드 리스트
url_list = ['categories/443129','categories/359743']
name_list = ['엉덩이탐정','츄츄']
campaign_list = []

dir_path = 'D:/backupGit/Crawling/'         # 크롤링 프로젝트 폴더경로

### 작성 구역 END ####--------------------------------------------------

crawling_work_path = dir_path + 'crawling_dir/' + big_cate_fileName
if not os.path.isdir(crawling_work_path) :
    os.makedirs(crawling_work_path)
merge_work_path = dir_path + 'merge_dir/' + big_cate_fileName
if not os.path.isdir(merge_work_path) :
    os.makedirs(merge_work_path)

def crawling() :
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(big_cate_fileName)
    wb.remove_sheet(wb['Sheet'])
    ws.append(['상품명', '가격', '대분류', '소분류', '재고'])

    i = 1

    while True:
        options = webdriver.ChromeOptions()

        options.add_argument('--headless')
        UserAgent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
        options.add_argument('user-agent=' + UserAgent)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        driver.get(url='https://www.coupang.com/np/'+ url_list[0] +'?page=' + str(i))
        time.sleep(3)

        try:
            product = driver.find_element(By.ID, 'productList')
            lis = product.find_elements(By.CLASS_NAME, 'baby-product')
            print('*' * 50 + ' ' + str(i) + 'Page Start!' + ' ' + '*' * 50)

            for li in lis:
                try:
                    product = li.find_element(By.CLASS_NAME, 'name').text
                    price = li.find_element(By.CLASS_NAME, 'price-value').text

                    print('Product: ' + product)
                    print('Price: ' + price)

                    ws.append([product, price, big_cate, middle_cate, str(10)])

                except Exception:
                    pass

            print('*' * 50 + ' ' + str(i) + 'Page End!' + ' ' + '*' * 50)
            time.sleep(2)
            i += 1
            driver.quit()

        except NoSuchElementException:
            wb.save(crawling_work_path + '/' + name_list[0] +'.xlsx')
            wb.close()
            url_list.pop(0)
            name_list.pop(0)

            sd.Beep(1000, 500)
            if len(url_list) == 0:
                exit(0)
            else:
                crawling()

def mergeFile():
    list = os.listdir(crawling_work_path)  # dir is your directory path
    print(list)

    f = open(merge_work_path + '/' + big_cate_fileName + '.csv', 'w', encoding='utf-8', newline='')
    wr = csv.writer(f)

    wr.writerow(["product_name", "price", "big_category", "small_category", "stock"])

    for file_path in list:
        wb_old = openpyxl.load_workbook(crawling_work_path + '/' + file_path)
        ws = wb_old.worksheets[0]
        for r in range(2, ws.max_row + 1):
            product_name = ws.cell(row=r, column=1).value
            price = ws.cell(row=r, column=2).value
            big_category = ws.cell(row=r, column=3).value
            small_category = ws.cell(row=r, column=4).value
            stock = ws.cell(row=r, column=5).value

            wr.writerow([product_name, price, big_category, small_category, stock])

    f.close()
    sd.Beep(1000, 500)

if crawling_flag == True:
    crawling()
else:
    mergeFile()
